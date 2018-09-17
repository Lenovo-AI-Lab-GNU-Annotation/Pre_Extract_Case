# coding: utf-8


"""
Created: Jian Chen 陈键 2018年07月04日
Last Modified: Jian Chen 陈键 2018年07月18日

该脚本读入原始excel表格，读取对话数据，并将关键case提取出来并在chatturns当中搜索找出对应数据，
将提取的数据储存到Template.xlsx中并另存为 output_name 文件。

运行方式：
python3 preExtractCase_v02.py <输入文件路径> <输出文件路径>
"""

import re,os,sys,codecs
from openpyxl import load_workbook
from collections import Counter

#%%

"""
输入，输出，template的路径和文件名的declaration

！！！！！！！！！！！！注意：template.xlsx一定要放在和代码同一个文件夹，否则会报错
"""

input_name = sys.argv[1]

output_name = sys.argv[2]

template = "template.xlsx"

#%%
"""
从文件名读取文件，从文件中读取sheet
"""
wb_template = load_workbook(filename = template)
wb_input = load_workbook(filename = input_name)

ws_rptChatVisit = wb_input['rptChatVisit']
ws_rptChatTurn = wb_input['rptChatTurn']

ws_extractData = wb_template['FacebookData']
ws_testerLog = wb_template['TesterLog']
ws_testerInfo = wb_template['TesterInfo']


#%%

#get tester IDs and Names.
#to seperate user data and tester data
"""
Tester ID 和 Name单独对待
"""
testerIDs = dict()

for row in ws_testerInfo.rows:
    if row[1].value == 'TesterID' or row[1].value is None:
        continue
    testerID = row[1].value
    testerName = row[0].value
    testerIDs[testerID]= testerName

    
#%%

#get the column names in template and input file

columnNameTemplate = dict()

for col in ws_extractData.columns:
    columnNameTemplate[col[1].value]=col[1].column

columnName_rptChatVisit = dict()

for col in ws_rptChatVisit.columns:
    columnName_rptChatVisit[col[1].value]=col[1].column

columnName_rptChatTurn = dict()

for col in ws_rptChatTurn.columns:
    columnName_rptChatTurn[col[1].value]=col[1].column

#单元测试   
for key in columnName_rptChatTurn:
    print (key,columnName_rptChatTurn[key])

#%%

"""
输入：chatlog
输出: 该chatlog中，user以text形式回复的最长一句，和以text形式回复的次数 (case,textNum)
"""
def extractCase(__chatlog__):
    case=''
    textNum=0
    chatlog_list=__chatlog__.split('\n')
    for line in chatlog_list:
        line=line.strip()
        if '\' disconnected ' in line:
            continue
        
        if 'Bot(' not in line and '(text)' in line:
            textNum+=1
            if len(case)<len(line):
                case=line
    
    return case,textNum



    

#%%    
"""
输入：chatlog
输出：在chatlog中搜索chatbot提供的按钮，有按钮时自动检索下一句，user按了helpful或者unhelpful则输出结果
否则输出 “No”

"""
def getHelpful(__chatlog__):
    
    helpful_info = "No"
    
    chatlog_list = re.split("\n",__chatlog__)
    
    for ind,item in enumerate(chatlog_list):
        
        if '[Helpful]' in item and len(chatlog_list)>ind+2:
        
           response = chatlog_list[ind+1]
           pos = response.find(']:')
            
           if pos == -1:
              print (i,ind,"?")
              continue
           
           response = response[pos+3:]
           
           if response == 'Helpful':
               helpful_info = 'Helpful'
           elif response == 'Unhelpful':
               helpful_info = 'Unhelpful'
    
    return helpful_info

#%%
"""
输入：chatlog
输出：在chatlog中搜索chatbot提供的按钮，有按钮时自动检索下一句，user按了helpful或者unhelpful则输出结果
否则输出 “No”
"""
def getIsToAgent(__chatlog__):

    isToAgent = "No"
    
    chatlog_list = re.split("\n",__chatlog__)
    
    for ind,item in enumerate(chatlog_list):
        
        if '[Chat with person]' in item and '(text)' not in item and len(chatlog_list)>ind+2:
            
           response = chatlog_list[ind+1]
           pos = response.find(']:')
    
           if pos == -1:
              print (i,ind,response)
              continue
           
           response = response[pos+3:]
           
           if response == 'Chat with person':
               isToAgent = 'ActiveToAgent'
    
    return isToAgent    

#%%
    
"""
主要脚本
"""

"""
from the input file get hash table
将chatturn页面的所有对话数据，及其对应的domain,intent,slot数据整理成一个哈希表
"""
cases = dict()

for i in range(2,ws_rptChatTurn.max_row):
    
    __case__ = ws_rptChatTurn['S'][i].value.strip()
    
    if len(__case__)==0:
        continue
    
    case_cut = __case__[:35]
    
    domain = ws_rptChatTurn['W'][i].value
    
    intent = ws_rptChatTurn['X'][i].value
    
    slot = ws_rptChatTurn['Y'][i].value
    
    if case_cut not in cases:
        
        cases[case_cut]=[domain,intent,slot]
    else:
        print (__case__,"同一case对应多个标签")
    
    if i%100 == 0:
        print ("哈希表正在生成","{}/{}".format(i,ws_rptChatTurn.max_row))


#%%
    
    
    
#%%    
#print (len(cases))
    
#%%
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

counter = 0


_row_ = 2     #用户数据
_row_tester = 2 #测试数据

cnt_TF = Counter() #测试用，和脚本无关

cnt_userids = dict() #测试用，和脚本无关

font_blue = Font(color ='000000FF')#如果dialog只有一项回复，将该输出case标蓝

visit_ids = []

"""
写入数据
注：如果如果只是列名字改了，在这里改就可以
"""

for i in range(2,ws_rptChatVisit.max_row):
                                                        #这里是输入文件列名
    serviceID = ws_rptChatVisit[columnName_rptChatVisit['serviceid']][i].value
    
    visit_ids.append(serviceID)
                                                        #这里是输入文件列名
    chatlog = ws_rptChatVisit[columnName_rptChatVisit['chatcontent']][i].value
                                                        #这里是输入文件列名
    userName = ws_rptChatVisit[columnName_rptChatVisit['username']][i].value
   
    case_ext, chatnum = extractCase(chatlog)
    
    cnt_TF[case_ext[:35] in cases] +=1
    
    if case_ext[:35] in cases:        
        tag_classified = cases[case_ext[:35]]
    else:
        tag_classified = ["","",""]
    
    isHelpful = getHelpful(chatlog)
    
    isToAgent = getIsToAgent(chatlog)
    
    _dom_ = tag_classified[0]
    
    _intent_ = tag_classified[1]
    
    _slot_ = tag_classified[2]
                                                    #这里是输入文件列名
    score = ws_rptChatVisit[columnName_rptChatVisit['score']][i].value
                                                    #这里是输入文件列名
    comment = ws_rptChatVisit[columnName_rptChatVisit['key_comments']][i].value
                                                    #这里是输入文件列名
    evals = ws_rptChatVisit[columnName_rptChatVisit['evaluate_words']][i].value
    
    if score == "[]":
        score = ""

    if comment == "[]":
        comment = ""
                                                #这里是输入文件列名
    date = ws_rptChatVisit[columnName_rptChatVisit['date']][i].value
    
    userId = ws_rptChatVisit[columnName_rptChatVisit['userid']][i].value.split("&&")[0]
      
    starttime = ws_rptChatVisit[columnName_rptChatVisit['start_time']][i].value
    
    endtime = ws_rptChatVisit[columnName_rptChatVisit['end_time']][i].value
    
    channel = ws_rptChatVisit[columnName_rptChatVisit['channel']][i].value
    
    if channel == 'facebook':
    
        cnt_userids[userId] = channel
    
    if userId in testerIDs:
        
        """
        如果是测试数据，写入测试人员页面TesterLog
        """
        _row_tester +=1
        print (_row_tester,userId,testerIDs[userId])
                                    #这里是template列名 如果template没改，这里不需要改
        ws_testerLog[columnNameTemplate['serviceId']+str(_row_tester)].value = serviceID
        ws_testerLog[columnNameTemplate['channel']+str(_row_tester)].value = channel
        ws_testerLog[columnNameTemplate['username']+str(_row_tester)].value = userName
        ws_testerLog[columnNameTemplate['chatlog']+str(_row_tester)].value = chatlog
        ws_testerLog[columnNameTemplate['start time']+str(_row_tester)].value = starttime
        ws_testerLog[columnNameTemplate['end time']+str(_row_tester)].value = endtime
        ws_testerLog[columnNameTemplate['date']+str(_row_tester)].value = date
        ws_testerLog[columnNameTemplate['Case']+str(_row_tester)].value = case_ext
        ws_testerLog[columnNameTemplate['isToAgent']+str(_row_tester)].value = isToAgent
        ws_testerLog[columnNameTemplate['helpful选项']+str(_row_tester)].value = isHelpful
        ws_testerLog[columnNameTemplate['Comments']+str(_row_tester)].value = comment
        ws_testerLog[columnNameTemplate['evaluate-input']+str(_row_tester)].value = evals
        ws_testerLog[columnNameTemplate['DomainResult']+str(_row_tester)].value = _dom_
        ws_testerLog[columnNameTemplate['IntentResult']+str(_row_tester)].value = _intent_
        ws_testerLog[columnNameTemplate['SlotResult']+str(_row_tester)].value = _slot_
        ws_testerLog[columnNameTemplate['UserID']+str(_row_tester)].value = userId
        
        if chatnum == 1:
            ws_testerLog[columnNameTemplate['Case']+str(_row_tester)].font = font_blue
            
    else:  
        """
        如果是用户数据，写入用户页面
        """
        _row_ +=1
        print ("row",_row_)
        ws_extractData[columnNameTemplate['serviceId']+str(_row_)].value = serviceID
        ws_extractData[columnNameTemplate['channel']+str(_row_)].value = channel
        ws_extractData[columnNameTemplate['username']+str(_row_)].value = userName
        ws_extractData[columnNameTemplate['chatlog']+str(_row_)].value = chatlog
        ws_extractData[columnNameTemplate['start time']+str(_row_)].value = starttime
        ws_extractData[columnNameTemplate['end time']+str(_row_)].value = endtime
        ws_extractData[columnNameTemplate['date']+str(_row_)].value = date
        ws_extractData[columnNameTemplate['Case']+str(_row_)].value = case_ext
        ws_extractData[columnNameTemplate['isToAgent']+str(_row_)].value = isToAgent
        ws_extractData[columnNameTemplate['helpful选项']+str(_row_)].value = isHelpful
        ws_extractData[columnNameTemplate['Comments']+str(_row_)].value = comment
        ws_extractData[columnNameTemplate['evaluate-input']+str(_row_)].value = evals
        ws_extractData[columnNameTemplate['DomainResult']+str(_row_)].value = _dom_
        ws_extractData[columnNameTemplate['IntentResult']+str(_row_)].value = _intent_
        ws_extractData[columnNameTemplate['SlotResult']+str(_row_)].value = _slot_
        ws_extractData[columnNameTemplate['UserID']+str(_row_)].value = userId
        
        if chatnum == 1:
            ws_extractData[columnNameTemplate['Case']+str(_row_tester)].font = font_blue        
        

print (cnt_TF)

wb_template.save(output_name)

#%%
"""
cnt_chatIDs = Counter()

for i in range(2,ws_rptChatTurn.max_row):
    
    cnt_chatIDs[ws_rptChatTurn['A'][i].value]+=1
    
print (cnt_chatIDs)
#%%
for item in  visit_ids:
    print (cnt_chatIDs[item])
    
#%%
    
ws_fbm = wb_input['fbmessage']

user_id_fb = Counter()

for i in range(2,ws_fbm.max_row):
    
    user_id_fb[ws_fbm['A'][i].value]+=1
    user_id_fb['all']+=1
    
print (user_id_fb)

#%%

#%%
"""



